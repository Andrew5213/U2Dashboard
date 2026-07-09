"""Helpers openpyxl para relatorios .xlsx que espelham a estrutura visual dos PDFs
(mesma paleta, mesmas secoes e ordem de report_service.py / rdo_pdf_service.py),
numa unica aba por relatorio, com secoes empilhadas verticalmente."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Paleta (mesma de report_service.py / rdo_pdf_service.py) ──────────────────
NAVY = "1B3A6B"
BLUE = "3B82F6"
GREEN = "16A34A"
RED = "DC2626"
AMBER = "D97706"
WHITE = "FFFFFF"
GRAY_50 = "F8FAFC"
GRAY_200 = "E2E8F0"
GRAY_600 = "475569"
DARK = "0F172A"

_THIN_BOTTOM = Border(bottom=Side(style="thin", color=GRAY_200))
_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
_SUBTITLE_FONT = Font(name="Calibri", size=10, color=WHITE)
_SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
_LABEL_FONT = Font(name="Calibri", size=8, bold=True, color=GRAY_600)
_VALUE_FONT = Font(name="Calibri", size=10, color=DARK)
_TABLE_HEADER_FONT = Font(name="Calibri", size=9, bold=True, color=GRAY_600)
_TABLE_CELL_FONT = Font(name="Calibri", size=9, color=DARK)
_KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
_KPI_LABEL_FONT = Font(name="Calibri", size=7, color=WHITE)
_ITALIC_FOOTER_FONT = Font(name="Calibri", size=8, italic=True, color=GRAY_600)


_INVALID_SHEET_CHARS = str.maketrans({c: "-" for c in r"\/?*[]:"})


def new_workbook(sheet_title: str) -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Relatorio").translate(_INVALID_SHEET_CHARS).strip("'")[:31] or "Relatorio"
    ws.sheet_view.showGridLines = False
    return wb, ws


def _fill_row(ws: Worksheet, row: int, n_cols: int, color: str) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)


def write_title(ws: Worksheet, row: int, title: str, subtitle: str = "", n_cols: int = 8) -> int:
    """Faixa de titulo (equivalente a capa do PDF). Retorna a proxima linha livre."""
    _fill_row(ws, row, n_cols, NAVY)
    ws.cell(row=row, column=1, value=title).font = _TITLE_FONT
    ws.row_dimensions[row].height = 26
    row += 1
    if subtitle:
        _fill_row(ws, row, n_cols, NAVY)
        ws.cell(row=row, column=1, value=subtitle).font = _SUBTITLE_FONT
        row += 1
    return row + 1


def write_section(ws: Worksheet, row: int, title: str, n_cols: int = 8) -> int:
    """Faixa de secao (equivalente ao section_header do PDF). Retorna a proxima linha livre."""
    _fill_row(ws, row, n_cols, NAVY)
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    return row + 1


def write_kv_rows(ws: Worksheet, row: int, pairs: list[tuple[str, object]], n_per_row: int = 3) -> int:
    """Pares label/valor em grade (equivalente ao kv_row do PDF). Retorna a proxima linha livre."""
    for i in range(0, len(pairs), n_per_row):
        chunk = pairs[i:i + n_per_row]
        for col, (label, _) in enumerate(chunk, start=1):
            c = ws.cell(row=row, column=col, value=label.upper())
            c.font = _LABEL_FONT
        row += 1
        for col, (_, value) in enumerate(chunk, start=1):
            c = ws.cell(row=row, column=col, value=value if value not in (None, "") else "-")
            c.font = _VALUE_FONT
        row += 1
    return row + 1


def write_table(
    ws: Worksheet, row: int, headers: list[str], rows: list[list[object]],
    col_widths: list[int] | None = None,
) -> int:
    """Tabela com cabecalho sombreado (equivalente a table_header/table_row do PDF).
    Retorna a proxima linha livre."""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _TABLE_HEADER_FONT
        c.fill = PatternFill("solid", fgColor=GRAY_50)
        c.border = _THIN_BOTTOM
    row += 1
    for r in rows:
        for col, val in enumerate(r, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _TABLE_CELL_FONT
            c.border = _THIN_BOTTOM
        row += 1
    if col_widths:
        for col, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w
    return row + 1


def write_kpis(ws: Worksheet, row: int, kpis: list[tuple[str, object, str]]) -> int:
    """Cartoes de KPI (equivalente ao kpi_box do PDF). kpis: [(label, value, hex_color)].
    Retorna a proxima linha livre."""
    for col, (label, value, color) in enumerate(kpis, start=1):
        vcell = ws.cell(row=row, column=col, value=value)
        vcell.font = _KPI_VALUE_FONT
        vcell.fill = PatternFill("solid", fgColor=color)
        vcell.alignment = Alignment(horizontal="center")
        lcell = ws.cell(row=row + 1, column=col, value=label.upper())
        lcell.font = _KPI_LABEL_FONT
        lcell.fill = PatternFill("solid", fgColor=color)
        lcell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 22
    return row + 3


def write_footer(ws: Worksheet, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = _ITALIC_FOOTER_FONT
    return row + 1
